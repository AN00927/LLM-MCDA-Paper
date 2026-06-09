#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
One-time migration: bring the consolidated workbook's three RAG sheets into the
*current* 90-scenario split (35 HVAC / 35 Appliance / 20 Shower) by overwriting
them, row-for-row, with the standalone RAG files (the pipeline's source of truth).

Why: ConsolidatedforSimaltaneousediting.xlsx still holds the OLD RAG partition
(38/39/29 = 106 scenarios) under an older column schema. rebuild_consolidated.py
re-derives Test = (masters not claimed by RAG), so a stale RAG set produces the
wrong Test/RAG partition. This script reseeds the consolidated RAG sheets from the
standalone files so a subsequent rebuild reproduces the current 195/90 split.

Masters are NOT touched here: master parameter values live only in the consolidated
workbook and are the source the standalone Test/RAG sheets were derived from. The
rebuild's audit + the reproduction diff are the gate that proves consistency.

Backs up the workbook first (timestamped), then verifies scenario_id counts.
"""

import os
import shutil
from datetime import datetime, timezone

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
WB_PATH = os.path.join(HERE, "ConsolidatedforSimaltaneousediting.xlsx")

# consolidated sheet name -> standalone source filename
SRC_MAP = {
    "HVACRAGScenarios": "HVACRagScenarios.xlsx",
    "ApplianceRAGScenarios": "ApplianceRAGScenarios.xlsx",
    "ShowerRAGScenarios": "ShowerRAGScenarios.xlsx",
}
EXPECTED_SCENARIOS = {"HVACRAGScenarios": 35, "ApplianceRAGScenarios": 35, "ShowerRAGScenarios": 20}


def backup():
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    bak = os.path.join(HERE, f"ConsolidatedforSimaltaneousediting_PREMIGRATION_{ts}.xlsx")
    shutil.copy2(WB_PATH, bak)
    if not os.path.exists(bak) or os.path.getsize(bak) == 0:
        raise RuntimeError("Backup failed or empty -> aborting")
    return bak


def read_standalone(fname):
    """Read the first worksheet of a standalone RAG file, preserving stored types."""
    wb = openpyxl.load_workbook(os.path.join(HERE, fname), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    # drop fully-blank trailing rows
    rows = [r for r in rows if any(v is not None and not (isinstance(v, str) and v.strip() == "") for v in r)]
    return rows


def main():
    bak = backup()
    print(f"[backup] {os.path.basename(bak)}")

    wb = openpyxl.load_workbook(WB_PATH)

    for sheet, fname in SRC_MAP.items():
        data = read_standalone(fname)
        ws = wb[sheet]
        # clear the entire old used range (old sheets are larger than the new split)
        for row in ws.iter_rows():
            for c in row:
                c.value = None
        # write headers + data from the standalone source verbatim
        for ri, rowvals in enumerate(data, start=1):
            for ci, v in enumerate(rowvals, start=1):
                ws.cell(ri, ci).value = v
        # scenario count (header is row 1; scenario_id is column 1 in all RAG sheets)
        sids = {r[0] for r in data[1:] if r and r[0] is not None}
        print(f"[migrate] {sheet} <- {fname}: {len(data) - 1} rows, {len(sids)} scenarios")

    wb.save(WB_PATH)

    # ---- verify ----
    wb2 = openpyxl.load_workbook(WB_PATH, read_only=True)
    ok = True
    for sheet, want in EXPECTED_SCENARIOS.items():
        ws = wb2[sheet]
        sids = set()
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue
            v = row[0].value
            if v is not None:
                sids.add(v)
        got = len(sids)
        status = "OK" if got == want else "MISMATCH"
        if got != want:
            ok = False
        print(f"[verify] {sheet}: {got} scenarios (want {want}) -> {status}")
    wb2.close()

    print("\n==================>", "MIGRATION OK" if ok else "MIGRATION COUNT MISMATCH", "<==================")
    raise SystemExit(0 if ok else 1)


if __name__ == "__main__":
    main()
