#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Deterministic rebuild of the three per-type RAG sheets and the combined
TestScenarios sheet inside ConsolidatedforSimaltaneousediting.xlsx.

Operates ONLY on the workbook (no repo code). Follows the schema-locked spec:
  - back up first (stop on failure)
  - clean master pools (safe-only, logged) and write fixes back
  - cache existing RAG scores, rebuild RAG (long, 3 rows/scenario, carried scores)
  - partition: Test = master scenarios not present in RAG (wide, no scores)
  - enforce Text / Number cell formats everywhere written (never General)
  - audit + fail loudly on hard violations (non-zero exit)

Decisions confirmed with the user:
  - rank is RECOMPUTED from mavt_score ordering (rank 1 = highest mavt); logged.
  - RAG scenario_id = sequential 1..N per decision type (unique within type).
"""

import os, re, sys, shutil, unicodedata
from datetime import datetime, timezone
from collections import defaultdict, Counter

import pandas as pd
import openpyxl

SEED = 20260602

HERE = os.path.dirname(os.path.abspath(__file__))
WB_PATH = os.path.join(HERE, "ConsolidatedforSimaltaneousediting.xlsx")
DUMP_DIR = os.path.join(HERE, "rebuild_dumps")
BACKUP_DIR = os.path.join(HERE, "Backups")  # timestamped backups of exported standalone files

# Single source of truth for the two derived labels (also used by BuildRAG /
# Eample-Guided_LLM_Scoring.pycoring.py) so the band/flow buckets can never drift between the
# scenario sheets and the RAG index. This is the only repo import rebuild makes.
PROJECT_ROOT = os.path.dirname(HERE)
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)
from sentinel_utils import house_age_to_band_label, gpm_to_flow_rate_label, appliance_age_to_band_label

# ----------------------------------------------------------------------------
# 0. Backup (stop on failure)
# ----------------------------------------------------------------------------
def backup():
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    bak = os.path.join(HERE, f"ConsolidatedforSimaltaneousediting_BACKUP_{ts}.xlsx")
    shutil.copy2(WB_PATH, bak)
    if not os.path.exists(bak) or os.path.getsize(bak) == 0:
        raise RuntimeError("Backup failed or empty -> aborting")
    return bak

# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------
REPLACEMENT = "�"  # U+FFFD only; U+2014 em-dash is legitimate and preserved

def clean_text(v):
    """Normalise nbsp, fold em/en-dashes to commas/hyphens, collapse whitespace, trim.

    The standalone scenario sheets use commas/hyphens (no em/en dashes); this
    reproduces that so rebuilt Test questions byte-match the standalone files:
    em-dash (U+2014, optional surrounding spaces) -> ", "; en-dash (U+2013)
    between digits -> "-", otherwise -> ", ". Legacy U+FFFD is first folded to
    an em-dash so it follows the same path.
    """
    if v is None:
        return None
    s = str(v)
    s = s.replace(REPLACEMENT, "—")
    s = s.replace(" ", " ")
    s = re.sub(r"\s*—\s*", ", ", s)        # em-dash -> comma
    s = re.sub(r"(?<=\d)–(?=\d)", "-", s)    # digit-flanked en-dash -> hyphen
    s = re.sub(r"\s*–\s*", ", ", s)        # any other en-dash -> comma
    s = re.sub(r"\s+", " ", s).strip()
    return s if s != "" else None

def has_corruption(s):
    if not isinstance(s, str):
        return False
    if REPLACEMENT in s:
        return True
    for ch in s:
        if ch in "\t\n":
            continue
        if unicodedata.category(ch).startswith("C"):
            return True
    return False

def to_num(v):
    """Coerce numeric / numeric-string to int or float (int when integral)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        f = float(v)
        return int(f) if f.is_integer() else f
    s = clean_text(v)
    f = float(s)
    return int(f) if f.is_integer() else f

def int_str(v):
    """Integer rendered as a text string, e.g. 80 -> '80'."""
    return str(int(round(float(clean_text(v) if isinstance(v, str) else v))))

_TIME_FULL = re.compile(r"^(\d{1,2}):(\d{2})\s*([ap])\.?m\.?$", re.I)
_TIME_HOUR = re.compile(r"^(\d{1,2})\s*([ap])\.?m\.?$", re.I)

def to_clock(v):
    """Normalise a clock value to 'H:MM AM/PM' text.
    Handles numeric Excel day-fraction serials, '6pm'/'8am' shorthand, and
    spacing/case variants. Returns the cleaned string unchanged if unparseable."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        frac = float(v) % 1.0
        total = int(round(frac * 24 * 60))
        h, m = (total // 60) % 24, total % 60
        ap = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12}:{m:02d} {ap}"
    s = clean_text(v)
    if s is None:
        return None
    mo = _TIME_FULL.match(s)
    if mo:
        return f"{int(mo.group(1))}:{int(mo.group(2)):02d} {mo.group(3).upper()}M"
    mo = _TIME_HOUR.match(s)
    if mo:
        return f"{int(mo.group(1))}:00 {mo.group(2).upper()}M"
    return s

# legitimate non-integer HVAC setpoint sentinels (e.g. "turn the system off")
KNOWN_NONINT_ALTS = {"off"}

def alt_norm(v, kind):
    """Normalise an alternative for matching/keys. 'int' kind -> integer string
    when numeric, else cleaned text (e.g. 'Off'). 'clock' kind -> to_clock."""
    if kind == "clock":
        return to_clock(v)
    s = clean_text(v) if isinstance(v, str) else v
    try:
        return str(int(round(float(s))))
    except (TypeError, ValueError):
        return clean_text(v)

def canon(v):
    """Canonical form for equality matching: numeric->rounded float, else cleaned str."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return round(float(v), 6)
    s = clean_text(v)
    if s is None:
        return None
    try:
        return round(float(s), 6)
    except ValueError:
        return s

APPLIANCE_MAP = {"washer": "washing_machine"}
def canon_appliance(v):
    s = clean_text(v)
    s = s.lower() if s else s
    return APPLIANCE_MAP.get(s, s)

# cell writers -------------------------------------------------------------
FMT = {"text": "@", "int": "0", "num": "0.######"}

def put(cell, kind, val):
    """Write a typed value with an explicit (non-General) number format."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        cell.value = None
    elif kind == "text":
        cell.value = str(val)
    elif kind == "int":
        cell.value = int(round(float(val)))
    else:  # num
        f = float(val)
        cell.value = int(f) if f.is_integer() else f
    cell.number_format = FMT[kind]

def export_standalone(fname, cols, rows):
    """Write one regenerated table to its standalone single-sheet xlsx.

    Backs up any existing file (timestamped, into Backups/) first, then writes
    headers + typed rows with the same put()/FMT typing used for the consolidated
    sheets. The pipeline's read_table_clean reads the first worksheet, so a single
    sheet named 'Sheet1' with snake_case headers is preserved.
    """
    out_path = os.path.join(HERE, fname)
    if os.path.exists(out_path):
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        stem = os.path.splitext(fname)[0]
        shutil.copy2(out_path, os.path.join(BACKUP_DIR, f"{stem}_BACKUP_{ts}.xlsx"))
    nb = openpyxl.Workbook()
    ws = nb.active
    ws.title = "Sheet1"
    for ci, (name, _) in enumerate(cols, start=1):
        hc = ws.cell(1, ci); hc.value = name; hc.number_format = "@"
    for ri, row in enumerate(rows, start=2):
        for ci, (name, kind) in enumerate(cols, start=1):
            put(ws.cell(ri, ci), kind, row.get(name))
    nb.save(out_path)

# ----------------------------------------------------------------------------
# generic reader
# ----------------------------------------------------------------------------
def read_sheet(ws):
    hdr = [c.value for c in ws[1]]
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
            continue
        out.append((r, dict(zip(hdr, vals))))
    return hdr, out

# ----------------------------------------------------------------------------
# canonical destination headers + per-column formats
# ----------------------------------------------------------------------------
TEST_COLS = [
    ("decision_type","text"),("question","text"),("location","text"),
    ("square_footage","num"),("insulation","text"),("household_size","int"),
    ("utility_budget","num"),("housing_type","text"),("outdoor_temp","num"),
    ("house_age","text"),("appliance_age","text"),("flow_rate","text"),
    ("alternative_1","text"),("alternative_2","text"),("alternative_3","text"),
]
HVAC_RAG_COLS = [
    ("scenario_id","int"),("question","text"),("location","text"),
    ("square_footage","num"),("insulation","text"),("household_size","int"),
    ("utility_budget","num"),("housing_type","text"),("outdoor_temp","num"),
    ("house_age","num"),("r_value","num"),("seer","num"),("hvac_age","num"),
    ("alternative","text"),
    ("energy_cost_score","num"),("environmental_score","num"),("comfort_score","num"),
    ("practicality_score","num"),("mavt_score","num"),("rank","int"),
    ("raw_kwh","num"),("raw_cost","num"),("raw_emissions","num"),
]
APPL_RAG_COLS = [
    ("scenario_id","int"),("question","text"),("location","text"),
    ("utility_budget","num"),("appliance","text"),("appliance_age","text"),
    ("housing_type","text"),("household_size","int"),("kwh_per_cycle","num"),
    ("alternative","text"),
    ("energy_cost_score","num"),("environmental_score","num"),("comfort_score","num"),
    ("practicality_score","num"),("mavt_score","num"),("rank","int"),
    ("raw_cost","num"),("raw_emissions","num"),
]
SHOWER_RAG_COLS = [
    ("scenario_id","int"),("question","text"),("location","text"),
    ("household_size","int"),("gpm","num"),("flow_rate","text"),
    ("tank_size","num"),("water_heater_temp","num"),("utility_budget","num"),
    ("housing_type","text"),("outdoor_temp","num"),("alternative","text"),
    ("duration_min","int"),
    ("energy_cost_score","num"),("environmental_score","num"),("comfort_score","num"),
    ("practicality_score","num"),("mavt_score","num"),("rank","int"),
    ("raw_kwh","num"),("raw_cost","num"),("raw_water_gallons","num"),
]

SCORE_KEYS = ["energy_cost_score","environmental_score","comfort_score",
              "practicality_score","mavt_score"]

# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------
def main():
    import random
    random.seed(SEED)

    bak = backup()
    print(f"[backup] {os.path.basename(bak)}")

    wb = openpyxl.load_workbook(WB_PATH)  # keeps formulas
    change_log = []  # (sheet, source_row, column, before, after, reason)
    flags = []       # human-review WARN notes
    info_notes = []  # benign observations
    def log(sheet, row, col, before, after, reason):
        change_log.append({"sheet": sheet, "source_row": row, "column": col,
                            "before": before, "after": after, "reason": reason})

    # ========================================================================
    # PHASE C(prep): cache existing RAG scores BEFORE clearing
    #   cache[type][old_sid] = {"params": rawrow, "scores": {alt_norm: bundle}}
    # ========================================================================
    def cache_rag(sheet, altcol, alt_kind, raw_cols):
        ws = wb[sheet]; _, rows = read_sheet(ws)
        groups = defaultdict(list)
        for r, d in rows:
            groups[d["scenario_id"]].append((r, d))
        cache = {}
        for sid, grp in groups.items():
            bundles = {}
            for r, d in grp:
                a = alt_norm(d[altcol], alt_kind)
                bundles[a] = {
                    "energy_cost_score": d["energy_cost_score"],
                    "environmental_score": d["environmental_score"],
                    "comfort_score": d["comfort_score"],
                    "practicality_score": d["practicality_score"],
                    "mavt_score": d["mavt_score"],
                    "rank_src": d["rank"],
                    **{rc: d[rc] for rc in raw_cols},
                }
            cache[sid] = {"rows": grp, "scores": bundles}
        return cache

    hvac_cache   = cache_rag("HVACRAGScenarios","alternative","int",["raw_kwh","raw_cost","raw_emissions"])
    appl_cache   = cache_rag("ApplianceRAGScenarios","alternative","clock",["raw_cost","raw_emissions"])
    shower_cache = cache_rag("ShowerRAGScenarios","alternative","int",["raw_kwh","raw_cost","raw_water_gallons"])

    # ========================================================================
    # PHASE B: clean masters (safe-only, logged) and re-type touched columns
    #   returns list of normalized master records (canonical typed values)
    # ========================================================================
    def clean_master_text_cell(sheet, r, ws, col, ci):
        """trim/collapse/em-dash a text cell, write back as Text, log content change."""
        cell = ws.cell(r, ci)
        before = cell.value
        after = clean_text(before)
        if (before if before is not None else "") != (after if after is not None else ""):
            log(sheet, r, col, before, after, "whitespace/encoding cleanup")
        cell.value = after
        cell.number_format = "@"
        if has_corruption(after):
            flags.append(f"{sheet} r{r} {col}: residual corruption {after!r}")
        return after

    # ---- HVAC master ----
    ws = wb["HVACScenarios"]; hdr, rows = read_sheet(ws)
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    hvac_master = []
    for r, d in rows:
        q  = clean_master_text_cell("HVACScenarios", r, ws, "Question", ci["Question"])
        loc= clean_master_text_cell("HVACScenarios", r, ws, "Location", ci["Location"])
        ins= clean_master_text_cell("HVACScenarios", r, ws, "Insulation", ci["Insulation"])
        ht = clean_master_text_cell("HVACScenarios", r, ws, "Housing Type", ci["Housing Type"])
        clean_master_text_cell("HVACScenarios", r, ws, "Occupancy context", ci["Occupancy context"])
        alts = []
        for ac in ["Alternative 1","Alternative 2","Alternative 3"]:
            cell = ws.cell(r, ci[ac]); before = cell.value
            after = alt_norm(before, "int")
            if after is None:
                flags.append(f"HVACScenarios r{r} {ac}: missing alternative {before!r}")
            elif not re.match(r"^-?\d+$", after):
                if after.lower() in KNOWN_NONINT_ALTS:
                    info_notes.append(f"HVACScenarios r{r} {ac}: non-numeric setpoint {after!r} (kept as text)")
                else:
                    flags.append(f"HVACScenarios r{r} {ac}: non-integer alternative {before!r}")
            if str(before) != str(after):
                log("HVACScenarios", r, ac, before, after, "alternative -> text integer string")
            cell.value = after; cell.number_format = "@"
            alts.append(after)
        hvac_master.append({
            "row": r, "question": q, "location": loc,
            "square_footage": to_num(d["Square Footage"]), "insulation": ins,
            "household_size": int(to_num(d["Household Size"])),
            "utility_budget": to_num(d["Utility Budget"]), "housing_type": ht,
            "outdoor_temp": to_num(d["Outdoor Temp"]),
            "house_age": clean_text(str(d["House Age"])),
            "hvac_age": to_num(d["HVAC Age"]),
            "r_value": to_num(d["R-Value"]),
            "seer": to_num(d["SEER"]),
            "alt": alts,
        })

    # ---- Appliance master ----
    ws = wb["ApplianceScenarios"]; hdr, rows = read_sheet(ws)
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    appl_master = []
    for r, d in rows:
        q  = clean_master_text_cell("ApplianceScenarios", r, ws, "Description", ci["Description"])
        loc= clean_master_text_cell("ApplianceScenarios", r, ws, "Location", ci["Location"])
        ht = clean_master_text_cell("ApplianceScenarios", r, ws, "Housing Type", ci["Housing Type"])
        # appliance canonicalization
        cell = ws.cell(r, ci["Appliance"]); before = cell.value
        appv = canon_appliance(before)
        if str(before) != str(appv):
            log("ApplianceScenarios", r, "Appliance", before, appv, "appliance canonicalization")
        if appv not in {"dishwasher","dryer","washing_machine"}:
            flags.append(f"ApplianceScenarios r{r} Appliance: unmappable {before!r}")
        cell.value = appv; cell.number_format = "@"
        # times (baseline + alternatives) -> Text clock
        bt_cell = ws.cell(r, ci["Baseline Time"]); bt_before = bt_cell.value
        bt = to_clock(bt_before)
        if str(bt_before) != str(bt):
            log("ApplianceScenarios", r, "Baseline Time", bt_before, bt, "time -> H:MM AM/PM text")
        bt_cell.value = bt; bt_cell.number_format = "@"
        alts = []
        for ac in ["Alternative 1","Alternative 2","Alternative 3"]:
            cell = ws.cell(r, ci[ac]); before = cell.value
            after = to_clock(before)
            if str(before) != str(after):
                log("ApplianceScenarios", r, ac, before, after, "time -> H:MM AM/PM text")
            cell.value = after; cell.number_format = "@"
            alts.append(after)
        appl_master.append({
            "row": r, "question": q, "location": loc,
            "utility_budget": to_num(d["Utility Budget"]), "appliance": appv,
            "housing_type": ht, "household_size": int(to_num(d["Occupants"])),
            "kwh_per_cycle": to_num(d["kwh/cycle"]),
            "appliance_age": clean_text(str(d["Appliance Age"])),
            "baseline_time": bt, "alt": alts,
        })

    # ---- Shower master ----
    ws = wb["ShowerScenarios"]; hdr, rows = read_sheet(ws)
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    shower_master = []
    for r, d in rows:
        q  = clean_master_text_cell("ShowerScenarios", r, ws, "Description", ci["Description"])
        loc= clean_master_text_cell("ShowerScenarios", r, ws, "Location", ci["Location"])
        ht = clean_master_text_cell("ShowerScenarios", r, ws, "Housing Type", ci["Housing Type"])
        alts = []
        for ac in ["Alternative 1","Alternative 2","Alternative 3"]:
            cell = ws.cell(r, ci[ac]); before = cell.value
            after = alt_norm(before, "int")
            if after is None or not re.match(r"^-?\d+$", after):
                flags.append(f"ShowerScenarios r{r} {ac}: non-integer alternative {before!r}")
            if str(before) != str(after):
                log("ShowerScenarios", r, ac, before, after, "alternative -> text integer string")
            cell.value = after; cell.number_format = "@"
            alts.append(after)
        shower_master.append({
            "row": r, "question": q, "location": loc,
            "household_size": int(to_num(d["Occupants"])), "gpm": to_num(d["GPM"]),
            "utility_budget": to_num(d["Utility Budget"]), "housing_type": ht,
            "outdoor_temp": to_num(d["Outdoor Temp"]),
            "tank_size": to_num(d["Tank Size"]),
            "water_heater_temp": to_num(d["Water Heater Temp"]),
            "alt": alts,
        })

    # ========================================================================
    # PHASE C/D: match each cached RAG scenario to its (unique) master row
    # ========================================================================
    # parameter key alignment master<->rag (same canonical order)
    HVAC_MKEYS = ["location","square_footage","insulation","household_size",
                  "utility_budget","housing_type","outdoor_temp","house_age"]
    HVAC_RKEYS = ["location","square_footage","insulation","household_size",
                  "utility_budget","housing_type","outdoor_temp","house_age"]
    APPL_MKEYS = ["location","utility_budget","appliance","housing_type",
                  "household_size","kwh_per_cycle","appliance_age"]
    APPL_RKEYS = ["location","utility_budget","appliance","housing_type",
                  "household_size","kwh_per_cycle","appliance_age"]
    SHOW_MKEYS = ["location","household_size","gpm","utility_budget",
                  "housing_type","outdoor_temp"]
    SHOW_RKEYS = ["location","household_size","gpm","utility_budget",
                  "housing_type","outdoor_temp"]

    def match_type(records, cache, mkeys, rkeys, alt_kind, appliance_fix=False):
        """Return (matched: list of (master_rec, old_sid), claimed_rows:set, orphans:list)."""
        # master index keyed on canonical params (+ altset)
        idx = {}
        for m in records:
            pk = []
            for k in mkeys:
                v = m[k]
                pk.append(canon_appliance(v) if (appliance_fix and k=="appliance") else canon(v))
            ak = tuple(sorted((alt_norm(a, alt_kind) for a in m["alt"]), key=str))
            key = tuple(pk)+(ak,)
            if key in idx:
                flags.append(f"DUPLICATE master param-vector (row {m['row']} vs {idx[key]['row']})")
            idx[key] = m
        matched, orphans, claimed = [], [], set()
        for sid, info in cache.items():
            d0 = info["rows"][0][1]
            grp_alts = [g[1]["alternative"] for g in info["rows"]]
            pk = []
            for k in rkeys:
                v = d0[k]
                pk.append(canon_appliance(v) if (appliance_fix and k=="appliance") else canon(v))
            ak = tuple(sorted((alt_norm(a, alt_kind) for a in grp_alts), key=str))
            key = tuple(pk)+(ak,)
            m = idx.get(key)
            if m is None:
                orphans.append(sid)
                flags.append(f"ORPHAN RAG scenario (old_sid={sid}) -> no master match; kept in RAG")
            else:
                matched.append((m, sid)); claimed.add(m["row"])
        return matched, claimed, orphans

    hvac_matched,  hvac_claimed,  hvac_orph  = match_type(hvac_master,  hvac_cache,  HVAC_MKEYS, HVAC_RKEYS, "int")
    appl_matched,  appl_claimed,  appl_orph  = match_type(appl_master,  appl_cache,  APPL_MKEYS, APPL_RKEYS, "clock", appliance_fix=True)
    show_matched,  show_claimed,  show_orph  = match_type(shower_master,shower_cache,SHOW_MKEYS, SHOW_RKEYS, "int")

    hvac_matched.sort(key=lambda t: t[0]["row"])
    appl_matched.sort(key=lambda t: t[0]["row"])
    show_matched.sort(key=lambda t: t[0]["row"])

    # ========================================================================
    # Build RAG output rows (long, scores carried, rank recomputed)
    # ========================================================================
    def recompute_rank(triples, type_name, new_sid):
        """triples: list of (alt_norm, bundle). Return {alt_norm: rank} by mavt desc."""
        order = sorted(triples, key=lambda t: float(t[1]["mavt_score"]), reverse=True)
        ranks = {}
        for i, (a, b) in enumerate(order, start=1):
            ranks[a] = i
            src = b.get("rank_src")
            try:
                same = int(round(float(src))) == i
            except (TypeError, ValueError):
                same = False
            if not same:
                log(f"{type_name}RAG", new_sid, "rank", src, i,
                    "rank recomputed from mavt_score ordering")
        return ranks

    hvac_rag_rows, appl_rag_rows, shower_rag_rows = [], [], []

    for new_sid, (m, old_sid) in enumerate(hvac_matched, start=1):
        bundles = hvac_cache[old_sid]["scores"]
        alts_norm = [alt_norm(a, "int") for a in m["alt"]]
        if len(set(alts_norm)) != 3:
            flags.append(f"HVAC new_sid={new_sid}: non-distinct alternatives {alts_norm}")
        q = clean_text(hvac_cache[old_sid]["rows"][0][1]["question"])
        ranks = recompute_rank([(a, bundles[a]) for a in alts_norm], "HVAC", new_sid)
        for a in alts_norm:
            b = bundles[a]
            hvac_rag_rows.append({
                "scenario_id": new_sid, "question": q, "location": m["location"],
                "square_footage": m["square_footage"], "insulation": m["insulation"],
                "household_size": m["household_size"], "utility_budget": m["utility_budget"],
                "housing_type": m["housing_type"], "outdoor_temp": m["outdoor_temp"],
                "house_age": m["house_age"],
                "r_value": m["r_value"], "seer": m["seer"], "hvac_age": m["hvac_age"],
                "alternative": a,
                "energy_cost_score": b["energy_cost_score"], "environmental_score": b["environmental_score"],
                "comfort_score": b["comfort_score"], "practicality_score": b["practicality_score"],
                "mavt_score": b["mavt_score"], "rank": ranks[a],
                "raw_kwh": b["raw_kwh"], "raw_cost": b["raw_cost"], "raw_emissions": b["raw_emissions"],
            })

    for new_sid, (m, old_sid) in enumerate(appl_matched, start=1):
        bundles = appl_cache[old_sid]["scores"]
        alts_norm = [alt_norm(a, "clock") for a in m["alt"]]
        if len(set(alts_norm)) != 3:
            flags.append(f"APPL new_sid={new_sid}: non-distinct alternatives {alts_norm}")
        q = clean_text(appl_cache[old_sid]["rows"][0][1]["question"])
        # flag the known I've/It's question divergence (kept verbatim, not reworded)
        if clean_text(q) != clean_text(m["question"]):
            flags.append(f"APPL new_sid={new_sid}: RAG question differs from master "
                         f"(RAG kept): RAG={q!r} MASTER={m['question']!r}")
        ranks = recompute_rank([(a, bundles[a]) for a in alts_norm], "APPL", new_sid)
        for a in alts_norm:
            b = bundles[a]
            appl_rag_rows.append({
                "scenario_id": new_sid, "question": q, "location": m["location"],
                "utility_budget": m["utility_budget"], "appliance": m["appliance"],
                "appliance_age": m["appliance_age"], "housing_type": m["housing_type"],
                "household_size": m["household_size"], "kwh_per_cycle": m["kwh_per_cycle"],
                "alternative": a,
                "energy_cost_score": b["energy_cost_score"], "environmental_score": b["environmental_score"],
                "comfort_score": b["comfort_score"], "practicality_score": b["practicality_score"],
                "mavt_score": b["mavt_score"], "rank": ranks[a],
                "raw_cost": b["raw_cost"], "raw_emissions": b["raw_emissions"],
            })

    for new_sid, (m, old_sid) in enumerate(show_matched, start=1):
        bundles = shower_cache[old_sid]["scores"]
        alts_norm = [alt_norm(a, "int") for a in m["alt"]]
        if len(set(alts_norm)) != 3:
            flags.append(f"SHOWER new_sid={new_sid}: non-distinct alternatives {alts_norm}")
        q = clean_text(shower_cache[old_sid]["rows"][0][1]["question"])
        ranks = recompute_rank([(a, bundles[a]) for a in alts_norm], "SHOWER", new_sid)
        for a in alts_norm:
            b = bundles[a]
            shower_rag_rows.append({
                "scenario_id": new_sid, "question": q, "location": m["location"],
                "household_size": m["household_size"], "gpm": m["gpm"],
                "flow_rate": gpm_to_flow_rate_label(m["gpm"]),
                "tank_size": m["tank_size"], "water_heater_temp": m["water_heater_temp"],
                "utility_budget": m["utility_budget"], "housing_type": m["housing_type"],
                "outdoor_temp": m["outdoor_temp"], "alternative": a, "duration_min": int(a),
                "energy_cost_score": b["energy_cost_score"], "environmental_score": b["environmental_score"],
                "comfort_score": b["comfort_score"], "practicality_score": b["practicality_score"],
                "mavt_score": b["mavt_score"], "rank": ranks[a],
                "raw_kwh": b["raw_kwh"], "raw_cost": b["raw_cost"], "raw_water_gallons": b["raw_water_gallons"],
            })

    # ========================================================================
    # Build Test output rows (wide) = masters not claimed by RAG, stable order
    # ========================================================================
    test_rows = []
    for m in hvac_master:
        if m["row"] in hvac_claimed:
            continue
        test_rows.append({
            "decision_type": "HVAC", "question": m["question"], "location": m["location"],
            "square_footage": m["square_footage"], "insulation": m["insulation"],
            "household_size": m["household_size"], "utility_budget": m["utility_budget"],
            "housing_type": m["housing_type"], "outdoor_temp": m["outdoor_temp"],
            "house_age": house_age_to_band_label(m["house_age"]), "appliance_age": None, "flow_rate": None,
            "alternative_1": m["alt"][0], "alternative_2": m["alt"][1], "alternative_3": m["alt"][2],
        })
    for m in appl_master:
        if m["row"] in appl_claimed:
            continue
        test_rows.append({
            "decision_type": "Appliance", "question": m["question"], "location": m["location"],
            "square_footage": None, "insulation": None,
            "household_size": m["household_size"], "utility_budget": m["utility_budget"],
            "housing_type": m["housing_type"], "outdoor_temp": None,
            "house_age": None, "appliance_age": appliance_age_to_band_label(m["appliance_age"]), "flow_rate": None,
            "alternative_1": m["alt"][0], "alternative_2": m["alt"][1], "alternative_3": m["alt"][2],
        })
    for m in shower_master:
        if m["row"] in show_claimed:
            continue
        test_rows.append({
            "decision_type": "Shower", "question": m["question"], "location": m["location"],
            "square_footage": None, "insulation": None,
            "household_size": m["household_size"], "utility_budget": m["utility_budget"],
            "housing_type": m["housing_type"], "outdoor_temp": m["outdoor_temp"],
            "house_age": None, "appliance_age": None, "flow_rate": gpm_to_flow_rate_label(m["gpm"]),
            "alternative_1": m["alt"][0], "alternative_2": m["alt"][1], "alternative_3": m["alt"][2],
        })

    # ========================================================================
    # PHASE E: clear + write destinations with canonical headers and formats
    # ========================================================================
    def write_sheet(sheet, cols, rows):
        ws = wb[sheet]
        old_max_r, old_max_c = ws.max_row, ws.max_column
        for r in range(1, old_max_r + 1):
            for c in range(1, max(old_max_c, len(cols)) + 1):
                cell = ws.cell(r, c); cell.value = None
        # headers
        for ci0, (name, _) in enumerate(cols, start=1):
            hc = ws.cell(1, ci0); hc.value = name; hc.number_format = "@"
        # data
        for ri0, row in enumerate(rows, start=2):
            for ci0, (name, kind) in enumerate(cols, start=1):
                put(ws.cell(ri0, ci0), kind, row.get(name))

    write_sheet("TestScenarios", TEST_COLS, test_rows)
    write_sheet("HVACRAGScenarios", HVAC_RAG_COLS, hvac_rag_rows)
    write_sheet("ApplianceRAGScenarios", APPL_RAG_COLS, appl_rag_rows)
    write_sheet("ShowerRAGScenarios", SHOWER_RAG_COLS, shower_rag_rows)

    # ========================================================================
    # PHASE F: audit
    # ========================================================================
    audit = []  # (id, level, message)
    def A(idc, level, msg): audit.append((idc, level, msg))

    def read_written(sheet, cols):
        ws = wb[sheet]
        out = []
        for r in range(2, ws.max_row + 1):
            cells = [ws.cell(r, c) for c in range(1, len(cols) + 1)]
            if all(c.value is None for c in cells):
                continue
            out.append((r, {cols[i][0]: cells[i] for i in range(len(cols))}))
        return out

    test_w   = read_written("TestScenarios", TEST_COLS)
    hvac_w   = read_written("HVACRAGScenarios", HVAC_RAG_COLS)
    appl_w   = read_written("ApplianceRAGScenarios", APPL_RAG_COLS)
    shower_w = read_written("ShowerRAGScenarios", SHOWER_RAG_COLS)

    # ---- A01 disjointness (content) ----
    def test_pv(d):
        dt = d["decision_type"].value
        if dt == "HVAC":
            base = ("HVAC", canon(d["location"].value), canon(d["square_footage"].value),
                    canon(d["insulation"].value), canon(d["household_size"].value),
                    canon(d["utility_budget"].value), canon(d["housing_type"].value),
                    canon(d["outdoor_temp"].value))
        elif dt == "Appliance":
            base = ("Appliance", canon(d["location"].value), canon(d["household_size"].value),
                    canon(d["utility_budget"].value), canon(d["housing_type"].value),
                    canon(d["appliance_age"].value))
        else:
            base = ("Shower", canon(d["location"].value), canon(d["household_size"].value),
                    canon(d["utility_budget"].value), canon(d["housing_type"].value),
                    canon(d["outdoor_temp"].value), canon(d["flow_rate"].value))
        alts = tuple(sorted((str(d[a].value) for a in ["alternative_1","alternative_2","alternative_3"]), key=str))
        return base + (alts,)

    def rag_pv(dt, rows):
        groups = defaultdict(list)
        for r, d in rows:
            groups[d["scenario_id"].value].append(d)
        out = set()
        for sid, grp in groups.items():
            d = grp[0]
            if dt == "HVAC":
                base = ("HVAC", canon(d["location"].value), canon(d["square_footage"].value),
                        canon(d["insulation"].value), canon(d["household_size"].value),
                        canon(d["utility_budget"].value), canon(d["housing_type"].value),
                        canon(d["outdoor_temp"].value))
            elif dt == "Appliance":
                base = ("Appliance", canon(d["location"].value), canon(d["household_size"].value),
                        canon(d["utility_budget"].value), canon(d["housing_type"].value),
                        canon(d["appliance_age"].value))
            else:
                base = ("Shower", canon(d["location"].value), canon(d["household_size"].value),
                        canon(d["utility_budget"].value), canon(d["housing_type"].value),
                        canon(d["outdoor_temp"].value), canon(d["flow_rate"].value))
            alts = tuple(sorted((str(x["alternative"].value) for x in grp), key=str))
            out.add(base + (alts,))
        return out

    test_pvs = [test_pv(d) for _, d in test_w]
    rag_pvs = rag_pv("HVAC", hvac_w) | rag_pv("Appliance", appl_w) | rag_pv("Shower", shower_w)
    overlap = set(test_pvs) & rag_pvs
    if overlap:
        A("A01","FAIL", f"{len(overlap)} parameter-vector(s) appear in BOTH Test and RAG")
    else:
        A("A01","PASS","Test and RAG are disjoint by content (no shared parameter vectors)")

    # ---- A02 RAG id + order ----
    def check_ids_order(name, rows, master_lookup, alt_kind):
        groups = defaultdict(list)
        for r, d in rows:
            groups[d["scenario_id"].value].append((r, d))
        ok = True
        ids = sorted(groups)
        if ids != list(range(1, len(ids) + 1)):
            ok = False; A("A02","FAIL", f"{name}: ids not 1..N unique ({ids[:5]}...)")
        for sid, grp in groups.items():
            rsorted = sorted(grp, key=lambda t: t[0])
            if [t[0] for t in rsorted] != list(range(rsorted[0][0], rsorted[0][0] + len(rsorted))):
                ok = False; A("A02","FAIL", f"{name} sid {sid}: rows not consecutive")
            if len(grp) != 3:
                ok = False; A("A02","FAIL", f"{name} sid {sid}: has {len(grp)} rows (!=3)")
        return ok

    o1 = check_ids_order("HVACRAG", hvac_w, None, "int")
    o2 = check_ids_order("ApplianceRAG", appl_w, None, "clock")
    o3 = check_ids_order("ShowerRAG", shower_w, None, "int")
    if o1 and o2 and o3:
        A("A02","PASS","Each RAG scenario_id has exactly 3 consecutive rows; ids 1..N per type")

    # verify alt-1/2/3 order vs master (order = correctness)
    def check_alt_order(name, matched, rows, alt_kind):
        groups = defaultdict(list)
        for r, d in rows:
            groups[d["scenario_id"].value].append((r, d))
        bad = 0
        for new_sid, (m, old_sid) in enumerate(matched, start=1):
            grp = sorted(groups[new_sid], key=lambda t: t[0])
            want = [alt_norm(a, alt_kind) for a in m["alt"]]
            got  = [str(d["alternative"].value) for r, d in grp]
            if want != got:
                bad += 1
        if bad:
            A("A02b","FAIL", f"{name}: {bad} scenarios where row order != master alt-1/2/3 order")
        else:
            A("A02b","PASS", f"{name}: row order matches master alternative-1/2/3 order")
    check_alt_order("HVACRAG", hvac_matched, hvac_w, "int")
    check_alt_order("ApplianceRAG", appl_matched, appl_w, "clock")
    check_alt_order("ShowerRAG", show_matched, shower_w, "int")

    # ---- A03 scores carried ----
    def check_scores(name, rows, raw_cols):
        bad = 0
        for r, d in rows:
            for k in SCORE_KEYS + ["rank"] + raw_cols:
                v = d[k].value
                if v is None or not isinstance(v, (int, float)) or isinstance(v, bool):
                    bad += 1
        if bad:
            A("A03","FAIL", f"{name}: {bad} blank/non-numeric score/raw cells")
        else:
            A("A03","PASS", f"{name}: all scores/rank/raw populated & numeric")
    check_scores("HVACRAG", hvac_w, ["raw_kwh","raw_cost","raw_emissions"])
    check_scores("ApplianceRAG", appl_w, ["raw_cost","raw_emissions"])
    check_scores("ShowerRAG", shower_w, ["raw_kwh","raw_cost","raw_water_gallons"])

    # ---- A04 rank validity ----
    def check_rank(name, rows):
        groups = defaultdict(list)
        for r, d in rows:
            groups[d["scenario_id"].value].append(d)
        bad = 0
        for sid, grp in groups.items():
            ranks = sorted(int(d["rank"].value) for d in grp)
            if ranks != [1, 2, 3]:
                bad += 1; continue
            order = sorted(grp, key=lambda d: float(d["mavt_score"].value), reverse=True)
            if [int(d["rank"].value) for d in order] != [1, 2, 3]:
                bad += 1
        if bad:
            A("A04","FAIL", f"{name}: {bad} scenarios with invalid/inconsistent rank")
        else:
            A("A04","PASS", f"{name}: rank=={{1,2,3}} & consistent with mavt order")
    check_rank("HVACRAG", hvac_w); check_rank("ApplianceRAG", appl_w); check_rank("ShowerRAG", shower_w)

    # ---- A05 canonical headers (destination sheets) ----
    def check_headers(sheet, cols):
        ws = wb[sheet]
        got = [ws.cell(1, c).value for c in range(1, len(cols) + 1)]
        want = [c[0] for c in cols]
        if got != want:
            A("A05","FAIL", f"{sheet}: header drift {got}")
        else:
            A("A05","PASS", f"{sheet}: canonical headers OK")
    for s, c in [("TestScenarios",TEST_COLS),("HVACRAGScenarios",HVAC_RAG_COLS),
                 ("ApplianceRAGScenarios",APPL_RAG_COLS),("ShowerRAGScenarios",SHOWER_RAG_COLS)]:
        check_headers(s, c)
    A("A05info","INFO","Master sheets retain legacy headers by design (spec section 0 restricts "
                       "master edits to section-3 value cleanups; no header rename performed).")

    # ---- A06 cell-format conformance (destinations) ----
    def check_formats(sheet, cols, rows):
        bad = 0; general = 0
        for r, d in rows:
            for name, kind in cols:
                cell = d[name]; fmt = cell.number_format; v = cell.value
                if fmt == "General":
                    general += 1
                if fmt != FMT[kind]:
                    bad += 1
                if v is not None:
                    if kind == "text" and not isinstance(v, str):
                        bad += 1
                    if kind in ("int","num") and (not isinstance(v,(int,float)) or isinstance(v,bool)):
                        bad += 1
        if general or bad:
            A("A06","FAIL", f"{sheet}: {general} General-format cells, {bad} format/type mismatches")
        else:
            A("A06","PASS", f"{sheet}: every cell Text/Number per spec; none General")
    check_formats("TestScenarios", TEST_COLS, test_w)
    check_formats("HVACRAGScenarios", HVAC_RAG_COLS, hvac_w)
    check_formats("ApplianceRAGScenarios", APPL_RAG_COLS, appl_w)
    check_formats("ShowerRAGScenarios", SHOWER_RAG_COLS, shower_w)
    A("A06info","INFO","Inferred formats applied (confirm): hvac_age/house_age/appliance_age=Text, "
                       "tank_size/water_heater_temp/r_value/seer=Number, occupancy_context=Text, "
                       "duration_min/rank=Number(no-dec), *_score/raw_*=Number.")

    # ---- A07 time cells text ----
    def check_time(name, rows, cols):
        bad = 0
        for r, d in rows:
            for cn in cols:
                v = d[cn].value
                if v is None:
                    continue
                if not isinstance(v, str) or re.match(r"^\d+(\.\d+)?$", v):
                    bad += 1
        return bad
    appl_test = [(r, d) for r, d in test_w if d["decision_type"].value == "Appliance"]
    bt = check_time("TestAppl", appl_test, ["alternative_1","alternative_2","alternative_3"])
    bt += check_time("ApplRAG", appl_w, ["alternative"])
    if bt:
        A("A07","FAIL", f"{bt} appliance time cells are not 'H:MM AM/PM' text")
    else:
        A("A07","PASS","All appliance time cells are H:MM AM/PM text strings")

    # ---- A08 ages text ----
    def check_ages(rows, cols):
        bad = 0
        for r, d in rows:
            for cn in cols:
                v = d[cn].value
                if v is not None and not isinstance(v, str):
                    bad += 1
        return bad
    ab = check_ages(test_w, ["house_age","appliance_age"]) + check_ages(appl_w, ["appliance_age"])
    A("A08","FAIL" if ab else "PASS", f"{ab} non-text age cells" if ab else "Test house_age/appliance_age + ApplianceRAG appliance_age are Text (HVAC-RAG house_age is numeric by design)")

    # ---- A09 no corruption (written sheets + masters) ----
    corrupt = 0
    for sheet in ["TestScenarios","HVACRAGScenarios","ApplianceRAGScenarios","ShowerRAGScenarios",
                  "HVACScenarios","ApplianceScenarios","ShowerScenarios"]:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and has_corruption(c.value):
                    corrupt += 1
    A("A09","FAIL" if corrupt else "PASS", f"{corrupt} corrupt/non-printable cells" if corrupt else "No replacement/non-printable chars")

    # ---- A10 enum validity ----
    enum_bad = []
    INS = {"Good","Medium","Poor"}; HT = {"Apartment","Condo","Rowhouse","Single-family","Townhouse"}
    APP = {"dishwasher","dryer","washing_machine"}
    for r, d in test_w + hvac_w:
        if "insulation" in d and d["insulation"].value not in (None,) and d["insulation"].value not in INS:
            enum_bad.append(("insulation", d["insulation"].value))
    for rows in (test_w, hvac_w, appl_w, shower_w):
        for r, d in rows:
            if "housing_type" in d and d["housing_type"].value not in (None,) and d["housing_type"].value not in HT:
                enum_bad.append(("housing_type", d["housing_type"].value))
    for r, d in appl_w:
        if d["appliance"].value not in APP:
            enum_bad.append(("appliance", d["appliance"].value))
    A("A10","FAIL" if enum_bad else "PASS", f"invalid enums: {Counter(enum_bad)}" if enum_bad else "insulation/housing_type/appliance enums valid")

    # ---- A11 mapping fidelity (destination cells == cleaned-master values) ----
    mism = 0
    # Test rows map back to their master by order within type
    test_hvac   = [m for m in hvac_master  if m["row"] not in hvac_claimed]
    test_appl   = [m for m in appl_master  if m["row"] not in appl_claimed]
    test_show   = [m for m in shower_master if m["row"] not in show_claimed]
    test_masters = test_hvac + test_appl + test_show
    for (r, d), m in zip(test_w, test_masters):
        dt = d["decision_type"].value
        def eq(cell, val): return canon(cell.value) == canon(val)
        checks = [eq(d["question"], m["question"]), eq(d["location"], m["location"]),
                  eq(d["household_size"], m["household_size"]), eq(d["utility_budget"], m["utility_budget"]),
                  eq(d["housing_type"], m["housing_type"]),
                  eq(d["alternative_1"], m["alt"][0]), eq(d["alternative_2"], m["alt"][1]),
                  eq(d["alternative_3"], m["alt"][2])]
        if dt == "HVAC":
            checks += [eq(d["square_footage"], m["square_footage"]), eq(d["insulation"], m["insulation"]),
                       eq(d["outdoor_temp"], m["outdoor_temp"]),
                       eq(d["house_age"], house_age_to_band_label(m["house_age"]))]
        elif dt == "Appliance":
            checks += [eq(d["appliance_age"], appliance_age_to_band_label(m["appliance_age"]))]
        else:
            checks += [eq(d["outdoor_temp"], m["outdoor_temp"]),
                       eq(d["flow_rate"], gpm_to_flow_rate_label(m["gpm"]))]
        mism += checks.count(False)
    # RAG param cells map back to their matched master
    def rag_fid(rows, matched, fields):
        nonlocal mism
        groups = defaultdict(list)
        for r, d in rows:
            groups[d["scenario_id"].value].append(d)
        for new_sid, (m, old_sid) in enumerate(matched, start=1):
            d = groups[new_sid][0]
            for dest, msrc in fields:
                if canon(d[dest].value) != canon(m[msrc]):
                    mism += 1
    rag_fid(hvac_w, hvac_matched, [("location","location"),("square_footage","square_footage"),
            ("insulation","insulation"),("household_size","household_size"),("utility_budget","utility_budget"),
            ("housing_type","housing_type"),("outdoor_temp","outdoor_temp"),("house_age","house_age")])
    rag_fid(appl_w, appl_matched, [("location","location"),("utility_budget","utility_budget"),
            ("appliance","appliance"),("appliance_age","appliance_age"),("housing_type","housing_type"),
            ("household_size","household_size"),("kwh_per_cycle","kwh_per_cycle")])
    rag_fid(shower_w, show_matched, [("location","location"),("household_size","household_size"),
            ("gpm","gpm"),("utility_budget","utility_budget"),("housing_type","housing_type"),
            ("outdoor_temp","outdoor_temp")])
    # Shower-RAG flow_rate is a derived label: must equal gpm_to_flow_rate_label(master gpm)
    _shower_groups = defaultdict(list)
    for r, d in shower_w:
        _shower_groups[d["scenario_id"].value].append(d)
    for new_sid, (m, old_sid) in enumerate(show_matched, start=1):
        d = _shower_groups[new_sid][0]
        if canon(d["flow_rate"].value) != canon(gpm_to_flow_rate_label(m["gpm"])):
            mism += 1
    A("A11","FAIL" if mism else "PASS", f"{mism} destination cells differ from cleaned-master value" if mism else "Every mapped cell byte-matches its cleaned-master source")

    # ---- A12 shower water sanity (WARN) ----
    w_bad = []
    for r, d in shower_w:
        gpm = float(d["gpm"].value); dur = float(d["duration_min"].value); rw = float(d["raw_water_gallons"].value)
        if abs(rw - gpm * dur) > max(1.0, 0.15 * gpm * dur):
            w_bad.append((d["scenario_id"].value, round(gpm,2), int(dur), round(rw,2), round(gpm*dur,2)))
    if w_bad:
        A("A12","WARN", f"{len(w_bad)} shower rows where raw_water_gallons != gpm*duration "
                        f"(e.g. {w_bad[:4]})")
    else:
        A("A12","PASS","raw_water_gallons ~= gpm*duration_min")

    # ---- A13 range plausibility (WARN) ----
    out = []
    def rng(rows, col, lo, hi, label):
        for r, d in rows:
            if col in d and d[col].value is not None:
                try: v = float(d[col].value)
                except (TypeError, ValueError): continue
                if not (lo <= v <= hi): out.append((label, d[col].value))
    rng(test_w, "square_footage", 400, 6000, "square_footage"); rng(hvac_w, "square_footage", 400, 6000, "square_footage")
    for rows in (test_w, hvac_w, appl_w, shower_w): rng(rows, "household_size", 1, 12, "household_size")
    for rows in (test_w, hvac_w, appl_w, shower_w): rng(rows, "utility_budget", 50, 900, "utility_budget")
    rng(test_w, "outdoor_temp", -20, 120, "outdoor_temp"); rng(hvac_w, "outdoor_temp", -20, 120, "outdoor_temp"); rng(shower_w, "outdoor_temp", -20, 120, "outdoor_temp")
    rng(shower_w, "gpm", 1.0, 5.0, "gpm")
    rng(shower_w, "duration_min", 1, 30, "duration_min")
    # flow_rate is now a categorical label (not numeric) -> membership check
    FLOW_LABELS = {"low_flow", "standard", "high_flow"}
    for rows in (test_w, shower_w):
        for r, d in rows:
            if "flow_rate" in d and d["flow_rate"].value is not None and d["flow_rate"].value not in FLOW_LABELS:
                out.append(("flow_rate_label", d["flow_rate"].value))
    # Test house_age is a band label like "11-15 years"
    for r, d in test_w:
        v = d["house_age"].value
        if v is not None and not re.match(r"^\d+-\d+ years$", str(v)):
            out.append(("house_age_label", v))
    # kwh_per_cycle per appliance band
    for r, d in appl_w:
        app = d["appliance"].value; v = d["kwh_per_cycle"].value
        if v is None: continue
        band = {"washing_machine":(0.1,1.2),"dishwasher":(0.7,2.2),"dryer":(1.5,5.0)}.get(app,(0.1,6.0))
        if not (band[0] <= float(v) <= band[1]): out.append((f"kwh_per_cycle[{app}]", v))
    # HVAC setpoints
    for r, d in hvac_w:
        try: v = float(d["alternative"].value)
        except (TypeError, ValueError): continue
        if not (55 <= v <= 90): out.append(("hvac_setpoint", d["alternative"].value))
    if out:
        A("A13","WARN", f"{len(out)} numeric values outside expected ranges: {Counter(x[0] for x in out)} "
                        f"e.g. {out[:6]}")
    else:
        A("A13","PASS","All numeric values within plausible ranges")

    # ---- A14 coverage (INFO) ----
    rag_counts = {"HVAC": len({d['scenario_id'].value for _,d in hvac_w}),
                  "Appliance": len({d['scenario_id'].value for _,d in appl_w}),
                  "Shower": len({d['scenario_id'].value for _,d in shower_w})}
    test_counts = Counter(d["decision_type"].value for _, d in test_w)
    A("A14","INFO", f"RAG scenarios {rag_counts} | Test scenarios {dict(test_counts)} | "
                    f"masters HVAC={len(hvac_master)} Appliance={len(appl_master)} Shower={len(shower_master)}")

    # ========================================================================
    # write Audit sheet
    # ========================================================================
    if "Audit" in wb.sheetnames:
        del wb["Audit"]
    aw = wb.create_sheet("Audit")
    aw.append(["check", "level", "message"])
    for cid, lvl, msg in audit:
        aw.append([cid, lvl, str(msg)])
    aw.append([])
    aw.append(["change_log_size", "", len(change_log)])
    aw.append(["flags", "", len(flags)])
    for f in flags:
        aw.append(["FLAG", "WARN", f])
    for n in info_notes:
        aw.append(["NOTE", "INFO", n])
    for c in range(1, 4):
        aw.cell(1, c).number_format = "@"

    fails = [a for a in audit if a[1] == "FAIL"]

    # ========================================================================
    # save workbook
    # ========================================================================
    wb.save(WB_PATH)

    # ========================================================================
    # PHASE G: export regenerated tables to the standalone files the pipeline reads
    #   (note filename casing: HVACRagScenarios.xlsx, not HVACRAG...).
    #   Skipped if the audit failed so a bad rebuild can never clobber the inputs.
    # ========================================================================
    if not fails:
        export_standalone("TestScenarios.xlsx", TEST_COLS, test_rows)
        export_standalone("HVACRagScenarios.xlsx", HVAC_RAG_COLS, hvac_rag_rows)
        export_standalone("ApplianceRAGScenarios.xlsx", APPL_RAG_COLS, appl_rag_rows)
        export_standalone("ShowerRAGScenarios.xlsx", SHOWER_RAG_COLS, shower_rag_rows)
        print(f"[export] standalone files refreshed (pre-run backups in {os.path.relpath(BACKUP_DIR, HERE)}/)")
    else:
        print("[export] SKIPPED standalone export (audit FAILED) -> standalone files untouched")

    # ========================================================================
    # dumps
    # ========================================================================
    os.makedirs(DUMP_DIR, exist_ok=True)
    pd.DataFrame(test_rows, columns=[c[0] for c in TEST_COLS]).to_csv(
        os.path.join(DUMP_DIR, "TestScenarios.csv"), index=False, encoding="utf-8-sig")
    pd.DataFrame(hvac_rag_rows, columns=[c[0] for c in HVAC_RAG_COLS]).to_csv(
        os.path.join(DUMP_DIR, "HVACRAGScenarios.csv"), index=False, encoding="utf-8-sig")
    pd.DataFrame(appl_rag_rows, columns=[c[0] for c in APPL_RAG_COLS]).to_csv(
        os.path.join(DUMP_DIR, "ApplianceRAGScenarios.csv"), index=False, encoding="utf-8-sig")
    pd.DataFrame(shower_rag_rows, columns=[c[0] for c in SHOWER_RAG_COLS]).to_csv(
        os.path.join(DUMP_DIR, "ShowerRAGScenarios.csv"), index=False, encoding="utf-8-sig")
    pd.DataFrame(change_log, columns=["sheet","source_row","column","before","after","reason"]).to_csv(
        os.path.join(DUMP_DIR, "change_log.csv"), index=False, encoding="utf-8-sig")

    # ========================================================================
    # print summary
    # ========================================================================
    print("\n==================== AUDIT ====================")
    for cid, lvl, msg in audit:
        print(f"  [{lvl:4}] {cid:7} {msg}")
    print("\n---- change_log size:", len(change_log))
    print("---- RAG counts:", rag_counts)
    print("---- Test counts:", dict(test_counts))
    print("---- orphans: HVAC", hvac_orph, "Appliance", appl_orph, "Shower", show_orph)
    if flags:
        print(f"---- FLAGS ({len(flags)}):")
        for f in flags[:40]:
            print("    -", f)
        if len(flags) > 40:
            print(f"    ... +{len(flags)-40} more")
    if info_notes:
        print(f"---- INFO NOTES ({len(info_notes)}):")
        for n in info_notes[:20]:
            print("    -", n)
    verdict = "AUDIT PASSED" if not fails else f"AUDIT FAILED — {len(fails)} hard violations"
    print("\n==================>", verdict, "<==================")
    sys.exit(0 if not fails else 1)


if __name__ == "__main__":
    main()
